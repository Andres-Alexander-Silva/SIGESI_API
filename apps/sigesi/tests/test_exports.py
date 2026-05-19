"""Smoke tests for /api/v1/reportes/exportar/<resource>/ xlsx endpoints."""
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from apps.sigesi.models import (
    Actividad,
    Indicador,
    MatriculaSemillero,
    MedicionIndicador,
    ProduccionAcademica,
)


BASE = '/api/v1/reportes/exportar/'
XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
ZIP_MAGIC = b'PK\x03\x04'  # xlsx is a zip


def _assert_xlsx_response(resp, expected_name_prefix):
    assert resp.status_code == 200, resp.content[:200]
    assert resp['Content-Type'] == XLSX_MIME
    assert expected_name_prefix in resp['Content-Disposition']
    assert resp.content[:4] == ZIP_MAGIC


def _load(resp):
    return load_workbook(BytesIO(resp.content))


# ---------------------------------------------------------------------------
# Happy-path admin tests for each of the 6 endpoints
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_admin_can_export_estudiantes(auth_client, admin_user, estudiante):
    resp = auth_client(admin_user).get(f'{BASE}estudiantes/')
    _assert_xlsx_response(resp, 'estudiantes_')
    wb = _load(resp)
    ws = wb.active
    assert ws['A1'].value == 'Cédula'
    # Header + at least 1 estudiante row
    assert ws.max_row >= 2


@pytest.mark.django_db
def test_admin_can_export_proyectos(auth_client, admin_user, proyecto):
    resp = auth_client(admin_user).get(f'{BASE}proyectos/')
    _assert_xlsx_response(resp, 'proyectos_')
    wb = _load(resp)
    ws = wb.active
    assert ws['A1'].value == 'Código'
    assert ws.max_row >= 2


@pytest.mark.django_db
def test_admin_can_export_avances(auth_client, admin_user, actividad, lider_estudiantil):
    from apps.sigesi.models import Evidencia
    Evidencia.objects.create(
        actividad=actividad, tipo='documento', titulo='Acta',
        descripcion='descripción', subido_por=lider_estudiantil,
    )
    resp = auth_client(admin_user).get(f'{BASE}avances/')
    _assert_xlsx_response(resp, 'avances_')
    wb = _load(resp)
    assert wb.active['A1'].value == 'Título'


@pytest.mark.django_db
def test_admin_can_export_producciones_academicas(
    auth_client, admin_user, proyecto, semillero_aprobado, lider_estudiantil
):
    p = ProduccionAcademica.objects.create(
        titulo='Paper', tipo='articulo', proyecto=proyecto, semillero=semillero_aprobado,
    )
    p.autores.set([lider_estudiantil])
    resp = auth_client(admin_user).get(f'{BASE}producciones-academicas/')
    _assert_xlsx_response(resp, 'producciones_academicas_')
    wb = _load(resp)
    assert wb.active['A1'].value == 'Título'


@pytest.mark.django_db
def test_admin_can_export_actividades(auth_client, admin_user, actividad):
    resp = auth_client(admin_user).get(f'{BASE}actividades/')
    _assert_xlsx_response(resp, 'actividades_')
    wb = _load(resp)
    assert wb.active['A1'].value == 'Proyecto'
    assert wb.active.max_row >= 2


@pytest.mark.django_db
def test_admin_can_export_indicadores(
    auth_client, admin_user, semillero_aprobado
):
    ind = Indicador.objects.create(
        nombre='Tasa X', descripcion='d', categoria='culminacion',
        unidad_medida='%',
    )
    MedicionIndicador.objects.create(
        indicador=ind, semillero=semillero_aprobado, semestre='2025-1', valor=42,
    )
    resp = auth_client(admin_user).get(f'{BASE}indicadores/')
    _assert_xlsx_response(resp, 'indicadores_')
    wb = _load(resp)
    assert wb.active['A1'].value == 'Indicador'
    assert wb.active.max_row >= 2


# ---------------------------------------------------------------------------
# Permission denials (one per endpoint family)
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_estudiante_cannot_export_anything(auth_client, estudiante):
    client = auth_client(estudiante)
    for path in ('estudiantes/', 'proyectos/', 'avances/',
                 'producciones-academicas/', 'actividades/', 'indicadores/'):
        resp = client.get(f'{BASE}{path}')
        assert resp.status_code == 403, f'{path} -> {resp.status_code}'


@pytest.mark.django_db
def test_lider_estudiantil_cannot_export(auth_client, lider_estudiantil):
    resp = auth_client(lider_estudiantil).get(f'{BASE}proyectos/')
    assert resp.status_code == 403


@pytest.mark.django_db
def test_unauthenticated_returns_401(api_client):
    resp = api_client.get(f'{BASE}proyectos/')
    assert resp.status_code == 401


# ---------------------------------------------------------------------------
# Filter / scope sanity
# ---------------------------------------------------------------------------

@pytest.mark.django_db
def test_estudiantes_filter_by_semillero_returns_one_matricula(
    auth_client, admin_user, estudiante, semillero_aprobado
):
    MatriculaSemillero.objects.create(
        estudiante=estudiante, semillero=semillero_aprobado, semestre='2025-1',
    )
    resp = auth_client(admin_user).get(
        f'{BASE}estudiantes/?semillero={semillero_aprobado.id}'
    )
    _assert_xlsx_response(resp, 'estudiantes_')
    ws = _load(resp).active
    # 1 header + 1 matrícula
    assert ws.max_row == 2
    # Estado matrícula column should be populated for this filter
    headers = [c.value for c in ws[1]]
    estado_col_idx = headers.index('Estado matrícula') + 1
    assert ws.cell(row=2, column=estado_col_idx).value == 'activa'


@pytest.mark.django_db
def test_director_semillero_only_sees_their_semillero_in_proyectos_export(
    auth_client, director_semillero, proyecto, programa
):
    """Build a sibling Semillero + Proyecto unrelated to director_semillero.
    The export must include only the project linked to his/her semillero.
    """
    from apps.sigesi.models import GrupoInvestigacion, Semillero, Proyecto

    # Another director, semillero and project
    from django.contrib.auth import get_user_model
    User = get_user_model()
    otro_dir = User.objects.create(
        username='otrodir', cedula='CC999999', correo_personal='otrodir@x.com',
        email='otrodir@inst.edu', roles=['director_semillero'], is_active=True,
    )
    otro_dir.set_password('x')
    otro_dir.save()

    otro_grupo = GrupoInvestigacion.objects.create(
        nombre='Otro Grupo', codigo='G99', fecha_creacion=date.today(),
        programa_academico=programa, director=otro_dir,
    )
    otro_sem = Semillero.objects.create(
        nombre='Otro Sem', codigo='S99', objetivo='x',
        fecha_creacion=date.today(),
        grupo_investigacion=otro_grupo, director=otro_dir,
        estado_aval=Semillero.EstadoAvalChoices.APROBADO,
    )
    otro_proy = Proyecto.objects.create(
        titulo='Otro Proy', codigo='POTRO', descripcion='d', objetivo_general='o',
        director=otro_dir,
    )
    otro_proy.semilleros.set([otro_sem])

    resp = auth_client(director_semillero).get(f'{BASE}proyectos/')
    _assert_xlsx_response(resp, 'proyectos_')
    ws = _load(resp).active
    headers = [c.value for c in ws[1]]
    codigo_col = headers.index('Código') + 1
    codigos_visibles = {ws.cell(row=r, column=codigo_col).value for r in range(2, ws.max_row + 1)}
    assert proyecto.codigo in codigos_visibles
    assert 'POTRO' not in codigos_visibles


@pytest.mark.django_db
def test_indicadores_no_filter_returns_all_in_scope(
    auth_client, admin_user, semillero_aprobado
):
    ind = Indicador.objects.create(
        nombre='X', descripcion='d', categoria='culminacion', unidad_medida='%',
    )
    MedicionIndicador.objects.create(
        indicador=ind, semillero=semillero_aprobado, semestre='2025-1', valor=10,
    )
    resp = auth_client(admin_user).get(f'{BASE}indicadores/')
    _assert_xlsx_response(resp, 'indicadores_')
    ws = _load(resp).active
    assert ws.max_row == 2  # header + 1
