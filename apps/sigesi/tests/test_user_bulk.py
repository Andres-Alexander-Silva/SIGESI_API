import io
import os

import openpyxl
import pytest
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.sigesi.models import User, ProgramaAcademico
from apps.sigesi.views.config.user_view import _cell_to_str

URL = '/api/v1/config/users/bulk-upload/formato/'
UPLOAD_URL = '/api/v1/config/users/bulk-upload/'

# Encabezados de la plantilla oficial (sin columna Username: se deriva del email).
HEADERS = [
    'Cédula', 'Nombres', 'Apellidos', 'Email Institucional',
    'Correo Personal', 'Teléfono', 'Roles', 'Código Estudiantil',
    'Programa Académico',
]

XLSX_CT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
PLANTILLA = os.path.join(settings.BASE_DIR, 'FORMATO DE REGISTRO DE ESTUDIANTES.xlsx')


def _as_upload(wb):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return SimpleUploadedFile('carga.xlsx', buffer.read(), content_type=XLSX_CT)


def _build_xlsx(rows, title_rows=0):
    """Hoja con `title_rows` filas de relleno arriba, luego HEADERS y los datos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(title_rows):
        ws.append(['—'])
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)
    return _as_upload(wb)


# ---------------------------------------------------------------- _cell_to_str

def test_cell_to_str_float_entero_sin_decimal():
    assert _cell_to_str(1090123456.0) == '1090123456'


def test_cell_to_str_int():
    assert _cell_to_str(1090123456) == '1090123456'


def test_cell_to_str_texto_intacto():
    assert _cell_to_str('  1090123456  ') == '1090123456'


# --------------------------------------------------------------- bulk ingest

@pytest.mark.django_db
def test_bulk_upload_cedula_numerica_no_arrastra_decimal(auth_client, admin_user):
    """Una cédula que llega como float de Excel se guarda sin el sufijo '.0'."""
    client = auth_client(admin_user)
    rows = [[1090123456.0, 'Ana', 'Pérez', 'ana@ufps.edu.co',
             'ana@gmail.com', 3001234567.0, 'estudiante', 172001.0, '']]
    resp = client.post(UPLOAD_URL, {'file': _build_xlsx(rows)}, format='multipart')

    assert resp.status_code == 200
    assert resp.data['creados'] == 1

    user = User.objects.get(correo_personal='ana@gmail.com')
    assert user.cedula == '1090123456'
    assert user.codigo_estudiantil == '172001'
    assert user.telefono == '3001234567'


@pytest.mark.django_db
def test_username_se_deriva_del_email_institucional(auth_client, admin_user):
    client = auth_client(admin_user)
    rows = [['1090123456', 'Pepito', 'Pérez', 'pepito@ufps.edu.co',
             'pepito@gmail.com', '', 'estudiante', '', '']]
    resp = client.post(UPLOAD_URL, {'file': _build_xlsx(rows)}, format='multipart')

    assert resp.status_code == 200
    user = User.objects.get(correo_personal='pepito@gmail.com')
    assert user.username == 'pepito'


@pytest.mark.django_db
def test_username_colision_se_sufija(auth_client, admin_user):
    """Dos correos con el mismo prefijo no rompen la restricción de unicidad."""
    client = auth_client(admin_user)
    User.objects.create(username='pepito', cedula='X1',
                        correo_personal='existente@x.com', email='otro@ufps.edu.co',
                        roles=['estudiante'])
    rows = [['111', 'Pepito', 'A', 'pepito@ufps.edu.co', 'p1@gmail.com', '', 'estudiante', '', '']]
    resp = client.post(UPLOAD_URL, {'file': _build_xlsx(rows)}, format='multipart')

    assert resp.status_code == 200
    user = User.objects.get(correo_personal='p1@gmail.com')
    assert user.username == 'pepito1'


@pytest.mark.django_db
def test_encabezados_no_en_primera_fila(auth_client, admin_user):
    """La fila de encabezados se detecta aunque haya filas de título arriba."""
    client = auth_client(admin_user)
    rows = [['900', 'Sara', 'Ruiz', 'sara@ufps.edu.co', 'sara@gmail.com', '', 'estudiante', '', '']]
    resp = client.post(UPLOAD_URL, {'file': _build_xlsx(rows, title_rows=2)}, format='multipart')

    assert resp.status_code == 200
    assert resp.data['creados'] == 1
    assert User.objects.filter(correo_personal='sara@gmail.com').exists()


@pytest.mark.django_db
def test_programa_academico_match_insensible_a_mayusculas(auth_client, admin_user):
    client = auth_client(admin_user)
    prog = ProgramaAcademico.objects.create(
        nombre='Ingeniería de Sistemas', codigo='IS', facultad='Ingeniería')
    rows = [['222', 'Leo', 'Díaz', 'leo@ufps.edu.co', 'leo@gmail.com', '',
             'estudiante', '', 'INGENIERÍA DE SISTEMAS']]
    resp = client.post(UPLOAD_URL, {'file': _build_xlsx(rows)}, format='multipart')

    assert resp.status_code == 200
    user = User.objects.get(correo_personal='leo@gmail.com')
    assert user.programa_academico_id == prog.id


@pytest.mark.django_db
def test_programa_academico_sin_match_queda_null(auth_client, admin_user):
    client = auth_client(admin_user)
    rows = [['333', 'Mia', 'Soto', 'mia@ufps.edu.co', 'mia@gmail.com', '',
             'estudiante', '', 'Programa Inexistente']]
    resp = client.post(UPLOAD_URL, {'file': _build_xlsx(rows)}, format='multipart')

    assert resp.status_code == 200
    user = User.objects.get(correo_personal='mia@gmail.com')
    assert user.programa_academico_id is None


@pytest.mark.django_db
def test_bulk_upload_con_plantilla_oficial(auth_client, admin_user):
    """Regresión: la plantilla real (título en filas 1-2, encabezados en la 3)
    se parsea correctamente y crea el usuario."""
    if not os.path.exists(PLANTILLA):
        pytest.skip("Plantilla oficial no encontrada en el repo.")

    client = auth_client(admin_user)
    wb = openpyxl.load_workbook(PLANTILLA)
    ws = wb.active
    # Encabezados en la fila 3 → datos en la fila 4 (mismo orden de columnas).
    ws.cell(row=4, column=1, value=1090123456.0)          # Cédula (float)
    ws.cell(row=4, column=2, value='Carlos')              # Nombres
    ws.cell(row=4, column=3, value='Niño')                # Apellidos
    ws.cell(row=4, column=4, value='carlos@ufps.edu.co')  # Email Institucional
    ws.cell(row=4, column=5, value='carlos@gmail.com')    # Correo Personal
    ws.cell(row=4, column=6, value=3009998877.0)          # Teléfono
    ws.cell(row=4, column=7, value='estudiante')          # Roles
    ws.cell(row=4, column=8, value=172999.0)              # Código Estudiantil
    ws.cell(row=4, column=9, value='')                    # Programa Académico

    resp = client.post(UPLOAD_URL, {'file': _as_upload(wb)}, format='multipart')

    assert resp.status_code == 200, resp.data
    assert resp.data['creados'] == 1
    user = User.objects.get(correo_personal='carlos@gmail.com')
    assert user.cedula == '1090123456'
    assert user.username == 'carlos'
    assert user.codigo_estudiantil == '172999'


@pytest.mark.django_db
def test_bulk_upload_requiere_rol_administrador(auth_client, estudiante):
    client = auth_client(estudiante)
    rows = [[123.0, 'X', 'Y', 'x@ufps.edu.co', 'x@gmail.com', '', 'estudiante', '', '']]
    resp = client.post(UPLOAD_URL, {'file': _build_xlsx(rows)}, format='multipart')
    assert resp.status_code == 403


@pytest.mark.django_db
def test_unauthenticated_cannot_download_formato(api_client):
    resp = api_client.get(URL)
    assert resp.status_code == 401


@pytest.mark.django_db
def test_authenticated_user_can_download_formato(auth_client, estudiante):
    client = auth_client(estudiante)
    resp = client.get(URL)
    assert resp.status_code == 200
    assert resp['Content-Type'] == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    assert 'attachment' in resp['Content-Disposition']
    assert 'FORMATO_DE_REGISTRO_DE_ESTUDIANTES.xlsx' in resp['Content-Disposition']
